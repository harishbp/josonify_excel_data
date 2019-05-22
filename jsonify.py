from openpyxl import load_workbook
import json
loc  = 'female_names.xlsx'
wb = load_workbook(loc)
ROW_START = 5
ROW_END = 105
sheet = wb['female1']


data = {
}
data['female'] = {}
for i in range(ROW_START,ROW_END):
    print(sheet.cell(row=i, column=2).value)


    data['female'][sheet.cell(row=i, column=2).value]={
        'gender':"female",
        'age':sheet.cell(row=i, column=5).value,
        'name':sheet.cell(row=i, column=4).value,
        'skin':sheet.cell(row=i, column=6).value
    }

with open('female_data.json', 'w') as outfile:
    json.dump(data, outfile)
